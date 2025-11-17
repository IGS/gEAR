import os
import sys
import tarfile
import typing
from abc import ABC, abstractmethod
from pathlib import Path

import anndata as ad
import numpy as np
import pandas as pd
import spatialdata as sd
import spatialdata_io as sdio
import xarray
from gear.utils import update_adata_with_ensembl_ids
from spatialdata.transformations import (
    Scale,
    Sequence,
    Translation,
    set_transformation,
)
from spatialdata_io.experimental import from_legacy_anndata, to_legacy_anndata

if typing.TYPE_CHECKING:
    from anndata import AnnData
    from spatialdata import SpatialData

class SpatialHandler(ABC):
    """
    Abstract base class for handling spatial transcriptomics data, providing a unified interface for managing and processing spatial and annotated data objects.

    This class defines properties and methods for:
    - Managing normalized table names.
    - Accessing and setting AnnData and SpatialData objects.
    - Querying metadata such as image presence, coordinate system, platform, and image names.
    - Reading spatial data files.
    - Filtering spatial data to image boundaries.
    - Converting spatial data to AnnData format.
    - Writing spatial data to Zarr and AnnData (H5AD) file formats.

    Subclasses must implement abstract properties and methods to specify platform-specific details and file reading logic.

    Attributes:
        NORMALIZED_TABLE_NAME (str): Default name for the normalized table.

    Properties:
        normalized_table_name (str): Name of the normalized table.
        adata (AnnData): The AnnData object associated with this handler.
        sdata (SpatialData): The SpatialData object associated with this handler.
        has_images (bool): Whether the handler has associated images (abstract).
        coordinate_system (str): The coordinate system used (abstract).
        platform (str): The name of the platform (abstract).
        img_name (str | None): The name of the associated image (abstract).

    Methods:
        process_file(filepath: str) -> "SpatialHandler": Reads and processes a spatial data file (abstract).
        merge_centroids_with_obs() -> pd.DataFrame: Converts spot data to a pandas DataFrame.
        convert_sdata_to_adata(include_images: bool | None = None, table_name=None) -> "SpatialHandler": Converts SpatialData to AnnData.
        extract_img() -> np.ndarray: Extracts an image as a NumPy array.
        scale_and_translate_sdata(set_to_zero=True, apply_scale=True) -> "SpatialHandler": Scales and translates spatial data.
        write_to_zarr(filepath: str | None =None, overwrite: bool = False) -> "SpatialHandler": Writes SpatialData to a Zarr file.
        write_to_h5ad(filepath: str | None=None) -> "SpatialHandler": Writes AnnData to an H5AD file.
    """

    NORMALIZED_TABLE_NAME = "table"

    def __init__(self):
        self._adata = None
        self._sdata = None
        self.originalFile = None

    @property
    def normalized_table_name(self):
        """
        Returns the name of the normalized table associated with this instance.

        Returns:
            str: The normalized table name.
        """
        return self.NORMALIZED_TABLE_NAME

    @property
    def adata(self) -> "AnnData":
        """
        Returns the AnnData object associated with this instance.

        Returns:
            AnnData: The underlying AnnData object.
        """
        return self._adata

    @adata.setter
    def adata(self, adata: "AnnData") -> None:
        """
        Sets the AnnData object for the instance.

        Parameters:
            adata (AnnData): The annotated data matrix to be assigned to the instance.
        """
        self._adata = adata

    @property
    def sdata(self) -> "SpatialData":
        """
        Returns the associated SpatialData object.

        Returns:
            SpatialData: The spatial data instance associated with this handler.
        """
        return self._sdata

    @sdata.setter
    def sdata(self, sdata: "SpatialData") -> None:
        """
        Sets the SpatialData object for the handler.

        Parameters:
            sdata (SpatialData): The SpatialData instance to assign to the handler.
        """
        self._sdata = sdata

    @property
    @abstractmethod
    def has_images(self) -> bool:
        """
        Checks if the spatial handler has associated images.

        Returns:
            bool: True if images are present, False otherwise.
        """
        pass

    @property
    @abstractmethod
    def coordinate_system(self) -> str:
        """
        Returns the coordinate system used by the spatial handler.

        Returns:
            str: A string representing the name or type of the coordinate system (e.g., 'WGS84', 'UTM', etc.).
        """
        pass

    @property
    @abstractmethod
    def region_id(self) -> str:
        """
        Returns the region ID used for spot data.

        This id should be found in your table, and serve as the index for the shape under "region_name".

        Returns:
            str: The region ID.
        """
        pass

    @property
    @abstractmethod
    def region_name(self) -> str:
        """
        Returns the name of the region used for spot data.

        A couple of ways to find this when viewing the SpatialData object:
        1. Look within the self.coordinate_system elements, and it is the name for the Shapes one. The 2nd dimension of .shape should be 2
        2. Within sdata.tables["table"].obs, the "region" column should have only 1 unique value that should match the Shape element name.

        Returns:
            str: The name of the region.
        """
        pass

    @property
    @abstractmethod
    def platform(self) -> str:
        """
        Returns the name of the platform as a string.

        Returns:
            str: The name of the platform.
        """
        pass

    @property
    @abstractmethod
    def img_name(self) -> str | None:
        """
        Returns the name of the image associated with the current instance.

        Returns:
            str | None: The image name if available, otherwise None.
        """
        pass

    @abstractmethod
    def process_file(self, filepath: str) -> "SpatialHandler":
        """
        Reads and processes a spatial data file from the given filepath.

        Args:
            filepath (str): The path to the spatial data file to be read.

        Returns:
            SpatialHandler: An instance of SpatialHandler containing the loaded spatial data.

        Raises:
            FileNotFoundError: If the specified file does not exist.
            IOError: If there is an error reading the file.
            ValueError: If the file format is invalid or unsupported.
        """
        pass

    def close_spatialdata(self) -> None:
        """
        Close and clean up resources associated with the object's spatial data (self.sdata).

        This method performs a best-effort teardown of common file-backed and on-disk
        data structures used for spatial AnnData objects and Zarr stores. It is
        intended to release file handles and other resources so that files can be
        moved/removed or the process can exit cleanly.

        Behavior:
        - Iterates over any tables in self.sdata.tables (if present). For each table,
            if it appears to be an AnnData object with .isbacked truthy, attempts to
            close its underlying file handle via adata.file.close().
        - Attempts to close any attributes on self.sdata named 'store', 'zarr_store',
            'zarr_group', or 'zarr_root' by calling a .close() method if present and
            callable.
        - Deletes the self.sdata attribute (best-effort) and runs garbage collection
            via gc.collect() to help free memory and release operating-system resources.

        Notes and guarantees:
        - This method swallows exceptions raised while closing individual resources.
            It is intentionally tolerant to partial failures so that other resources may
            still be released.
        - It does not remove or modify on-disk data; it only attempts to close open
            handles and drop in-memory references.
        - The caller should drop any other references to the same AnnData / store
            objects (e.g., local variables) to allow them to be fully freed.
        - The operation is idempotent in intent (calling it multiple times should not
            raise), but thread-safety is not guaranteed. Avoid concurrent calls from
            multiple threads.

        Parameters:
        - None

        Returns:
        - None

        NOTE: Copilot-generated function

        Example:
                # Best practice: drop other references to sdata and then call:
                handler.close_spatialdata()
                # If needed, reopen the data later by re-reading from disk/storage.
        """
        import gc

        # Close backed AnnData tables (matches pattern used in your code)
        for tbl in getattr(self.sdata, "tables", {}).values():
            try:
                adata = tbl
                if getattr(adata, "isbacked", False):
                    try:
                        adata.file.close()
                    except Exception:
                        pass
            except Exception:
                pass

        # Try to close zarr/store objects if present
        for name in ("store", "zarr_store", "zarr_group", "zarr_root"):
            store = getattr(self.sdata, name, None)
            if store is not None:
                close_fn = getattr(store, "close", None)
                if callable(close_fn):
                    try:
                        close_fn()
                    except Exception:
                        pass

        # Drop references and collect
        try:
            # caller should also drop their reference to sdata
            self._sdata = None
        except Exception:
            pass
        gc.collect()

    def convert_sdata_to_adata(self, include_images: bool | None = None, table_name=None) -> "SpatialHandler":
        """
        Converts the internal spatial data object (`sdata`) to an AnnData object and assigns it to `self.adata`.

        Parameters:
            include_images (bool | None, optional): Whether to include image data in the conversion. If None, defaults to `self.has_images`.
            table_name (str, optional): The name of the table to use for conversion. If None, defaults to `self.NORMALIZED_TABLE_NAME`.

        Returns:
            SpatialHandler: The current instance with the `adata` attribute set.

        Raises:
            Exception: If `self.sdata` is None or if an error occurs during conversion.

        NOTE: This is a slow process that can be slower if include_images=True. It is better to read either an h5ad directly
        or read AnnData from SpatialData.tables["table"]
        """
        if self.sdata is None:
            raise Exception("No spatial data object present to convert to AnnData object.")

        if include_images is None:
            include_images = self.has_images

            # TODO: If sdata.image has a transformation applied, we need to undo it, as to_legacy_anndata will apply it again.


        # Generally everything should already be converted to the normalized table name
        if table_name is None:
            table_name = self.NORMALIZED_TABLE_NAME

        try:
            # table name should already be "table" for Visium
            adata = to_legacy_anndata(self.sdata, include_images=include_images, coordinate_system=self.coordinate_system, table_name=table_name)
        except Exception as err:
            print(str(err))
            raise Exception("Error occurred while converting spatial data object to AnnData object: " + str(err))
        self.adata = adata
        return self

    def extract_img(self) -> np.ndarray:
        """
        Extracts an image from the spatial data object and returns it as a NumPy array in (y, x, c) format.

        The method retrieves the image specified by `self.img_name` from `self.sdata.images`. If the image is a
        `xarray.DataTree`, it extracts the base pyramid level. The image is then converted from an xarray DataArray
        to a NumPy array and its axes are rearranged from (c, y, x) to (y, x, c).

        Returns:
            np.ndarray: The extracted image as a NumPy array in (y, x, c) format.

        Raises:
            Exception: If `self.img_name` is not specified.
        """

        if not self.img_name:
            raise Exception("No image name specified for conversion to 2D array.")

        img = self.sdata.images[self.img_name]
        if isinstance(img, xarray.DataTree):
            img = sd.get_pyramid_levels(img, n=0)

        # Convert xarray DataArray to numpy array
        img = img.to_numpy()

        # Currently the image is in (c, y, x) format, and needs to be converted to (y, x, c) format
        return np.moveaxis(img, 0, -1)

    def merge_centroids_with_obs(self) -> "SpatialHandler":
        """
        Merges spatial centroid coordinates with the observation table of the spatial dataset.

        This method retrieves centroid coordinates for the specified region and coordinate system,
        renames the coordinate columns to 'spatial1' and 'spatial2', and merges them with the
        observation DataFrame (`obs`) of the main data table. The merge is performed as an inner join
        on the region identifier to ensure only matching observations are retained. Handles both
        pandas and Dask DataFrames by computing them if necessary.

        Raises:
            ValueError: If centroid extraction fails.

        Returns:
            SpatialHandler: The current instance with updated observation data including centroid coordinates.
        """

        centroids_df =  sd.get_centroids(self.sdata[self.region_name], coordinate_system=self.coordinate_system)

        if centroids_df is None:
            raise ValueError("Could not extract spatial observation locations")

        # Add spatial coords. Not all locations may be present in adata after filtering
        centroids_df = centroids_df.rename(columns={"x": "spatial1", "y": "spatial2"})
        dataframe = self.sdata.tables["table"].obs

        # Compute the dataframe if it's a Dask dataframe (since we cannot merge into a Dask dataframe)
        if hasattr(dataframe, "compute"):
            dataframe = dataframe.compute()
        if hasattr(centroids_df, "compute"):
            centroids_df = centroids_df.compute()

        # Add the centroid info to the AnnData table. Inner join in case location ID does not exist in observation
        self.sdata.tables["table"].obs = dataframe.merge(centroids_df, on=self.region_id, how="inner")
        return self

    def scale_and_translate_sdata(self, set_to_zero=True, apply_scale=True) -> "SpatialHandler":
        """
        Filters the spatial data (`sdata`) to fit within the boundaries of the high-resolution image,
        optionally translating the coordinates to start at zero and scaling the data so that the longest
        dimension does not exceed 2000 pixels.

        Args:
            set_to_zero (bool, optional): If True, translates the data so that both x and y coordinates start at 0.
                Defaults to True.
            apply_scale (bool, optional): If True, scales the data so that the longest dimension is no more than 2000 pixels.
                Defaults to True.

        Returns:
            SpatialHandler: The updated SpatialHandler instance with filtered spatial data.
        """
        # Filter to only the hires image boundaries
        if not self.img_name:
            return self

        if self.platform == "visium":
            # Visium data is already in image space, so no need to scale or translate
            # SAdkins - honestly, not sure how to fix these yet as the shapes seem to translate wildly off of the image
            return self

        # Extent should be based on the image data, in case the observation data bleeds past the image
        img_extent = sd.get_extent(self.sdata[self.img_name], coordinate_system=self.coordinate_system)

        MAX_X = 2000
        # SAFETY CHECK: If img_extent width is less than or equal to MAX_X, just return.
        # Would rather retain extra data and manually process than lose it.
        img_x_width = img_extent["x"][1] - img_extent["x"][0]
        if img_x_width <= MAX_X:
            return self

        sdata: "SpatialData" = sd.bounding_box_query(self.sdata,
                axes=("x", "y"),
                min_coordinate=[img_extent["x"][0], img_extent["y"][0]],
                max_coordinate=[img_extent["x"][1], img_extent["y"][1]],
                target_coordinate_system=self.coordinate_system,
                filter_table=True,
                ) # type: ignore

        # Get the region extent based on the filtered data
        region_extent = sd.get_extent(sdata[self.region_name], coordinate_system=self.coordinate_system)

        # NOTE: The "set_transformation" only modifies metadata about a transformation that is used in other spatialdata functions.
        # To actually apply the transformation run "spatialdata.transform()" or a related function

        # Perform a translation to the coordinate system so that both x and y start at 0
        # While it is possible to save a "transform" (and time) by combining the two in a Sequence transformation,
        # this is necessary to ensure the scaling is done correctly based on the updated coordinates.
        if set_to_zero:
            translation = Translation([-region_extent["x"][0], -region_extent["y"][0]], axes=["x", "y"])
            set_transformation(sdata[self.region_name], translation, to_coordinate_system=self.coordinate_system)
            sdata.transform_element_to_coordinate_system(self.region_name, target_coordinate_system=self.coordinate_system)

            # Recalculate the region extent after translation, so that scaling is correct
            region_extent = sd.get_extent(sdata[self.region_name], coordinate_system=self.coordinate_system)

        # Ensure the longest dimension is no more than 2000 pixels. This is the "legacy_to_anndata" hires pixel limit.
        if apply_scale:
            region_x_width = region_extent["x"][1] - region_extent["x"][0]

            if region_x_width > MAX_X:
                scale_factor_x = MAX_X / region_x_width

                scale = Scale([scale_factor_x, scale_factor_x], axes=("x", "y"))
                set_transformation(sdata[self.region_name], scale, to_coordinate_system=self.coordinate_system)
                sdata.transform_element_to_coordinate_system(self.region_name, target_coordinate_system=self.coordinate_system)

            # Rasterize the image if it exceeds the max width
            img_x_width = img_extent["x"][1] - img_extent["x"][0]
            if img_x_width > MAX_X:
                rasterized = sd.rasterize(
                        sdata[self.img_name],
                        axes=("x", "y"),
                        min_coordinate=[img_extent[ax][0] for ax in ("x", "y")],
                        max_coordinate=[img_extent[ax][1] for ax in ("x", "y")],
                        target_coordinate_system=self.coordinate_system,
                        target_unit_to_pixels=None,
                        target_width=MAX_X,
                        target_height=None,
                        target_depth=None,
                        return_regions_as_labels=True,
                    )
                sdata[self.img_name] = rasterized

        self.sdata = sdata
        return self

    def standardize_sdata(self) -> "SpatialHandler":
        """
        Normalize and preprocess the spatial dataset in self.sdata.tables["table"].

        This method performs an in-place spatial and single-cell style preprocessing pipeline on the
        SpatialHandler's SpatialData (self.sdata). It executes the following high-level steps in order:

        1. Subset the SpatialData
            - Calls self.subset_sdata() to apply any configured subsetting filters to the spatial object.

        2. Scale and translate coordinates
            - Calls self.scale_and_translate_sdata() to convert/adjust coordinate systems so that spatial
            annotations align with image space as required.

        3. Merge polygon centroids into observations
            - Calls self.merge_centroids_with_obs() to compute centroids from per-observation polygon shapes
            and merge those coordinates into the observation table.

        4. Single-cell preprocessing on the AnnData table
            - Loads the AnnData stored at self.sdata.tables["table"] and runs a Scanpy workflow:
                - sc.pp.normalize_total on the AnnData (inplace)
                - sc.pp.log1p
                - adata.var_names_make_unique()
                - sc.pp.highly_variable_genes(adata, n_top_genes=2000)
                - sc.pp.pca(adata)
                - sc.pp.neighbors(adata)
                - sc.tl.umap(adata)
            - The processed AnnData replaces the original in self.sdata.tables["table"].

        Returns:
            SpatialHandler: Returns self to allow method chaining.

        Side effects:
            - Modifies self.sdata in-place, including self.sdata.tables["table"] (AnnData) and any spatial
                coordinate fields produced by the called helper methods.
            - Requires that self.sdata.tables["table"] exists and is a valid AnnData object.
            - Requires the scanpy library to be available; an ImportError will occur if scanpy is not installed.

        Notes:
            - The number of highly variable genes is fixed to 2000 in this method. Adjustments require
            changing the implementation.
            - This method assumes that polygon shapes for observations are available so centroids can be
            computed and merged into observation metadata.
            - Intended for workflows that combine image-based spatial annotations with single-cell-style
            expression preprocessing.
        """
        obs = self.sdata.tables["table"].obs

        if not ("spatial1" in obs.columns and "spatial2" in obs.columns):
            self.subset_sdata()
            self.scale_and_translate_sdata()

            # The SpatialData object table should have coordinates, but they are not translated into the image space
            # Each observation has an associated polygon "shape" in the image space, and we can get the centroid of that shape
            self.merge_centroids_with_obs()

        # Run the single-cell workbench steps on the spatial_obj.tables["table"] (AnnData object) using default parameters
        adata = self.sdata.tables["table"]

        if "X_umap" in adata.obsm.keys():
            return self

        import scanpy as sc

        sc.pp.normalize_total(adata, inplace=True)
        sc.pp.log1p(adata)
        adata.var_names_make_unique()

        # Add qc-metrics (so we can filter on them later if desired)
        sc.pp.calculate_qc_metrics(adata, log1p=False, percent_top=None, inplace=True)

        sc.pp.highly_variable_genes(adata, n_top_genes=2000)
        sc.pp.pca(adata)
        sc.pp.neighbors(adata)
        sc.tl.umap(adata)
        self.sdata.tables["table"] = adata
        return self

    def subset_sdata(self) -> "SpatialHandler":
        """
        Subsets the spatial data (`sdata`) to include only specific elements and updates the instance's `sdata` attribute.

        The subset includes the normalized table, region name, and image name. The method also filters tables within the data.

        Returns:
            SpatialHandler: The instance with the updated, subsetted `sdata`.
        """
        subset_elements = [self.NORMALIZED_TABLE_NAME, self.region_name, self.img_name]
        self.sdata = self.sdata.subset(subset_elements, filter_tables=True)
        return self

    def write_to_zarr(self, filepath: str | None =None, overwrite: bool = False) -> "SpatialHandler":
        """
        Writes the spatial data object to a Zarr file at the specified file path.

        This method performs the following steps:
        - Checks that a spatial data object (`self.sdata`) is present.
        - Ensures a destination file path is provided.
        - Adds the platform type as metadata to the normalized table.
        - Converts the data matrix to float type for compatibility.
        - Attempts to write the spatial data to the specified file path.
        - If an error occurs during writing, removes any partially created directory.

        Args:
            filepath (str | None): The destination file path where the Zarr file will be written.
            overwrite (bool): Whether to overwrite the file if it already exists. Defaults to False.

        Returns:
            SpatialHandler: The current instance of the SpatialHandler.

        Raises:
            Exception: If no spatial data object is present, no file path is provided, or an error occurs during writing.
        """
        if self.sdata is None:
            raise Exception("No spatial data object present to write to file.")
        if filepath is None:
            raise Exception("No destination file path given. Provide one to write file.")
        try:
            # Add platform type as metadata. Can use downstream
            self.sdata.tables[self.NORMALIZED_TABLE_NAME].uns["platform"] = self.platform

            # See https://github.com/pydata/xarray/issues/3476#issuecomment-1115045538 for why we need to convert to unicode
            self.sdata.tables[self.NORMALIZED_TABLE_NAME].X = self.sdata.tables[self.NORMALIZED_TABLE_NAME].X.astype("float")

            # Will fail if file already exists
            self.sdata.write(file_path=filepath, overwrite=overwrite)
        except Exception as err:
            # remove the directory if it was created
            if os.path.exists(filepath):
                import shutil
                shutil.rmtree(filepath)
            raise Exception("Error occurred while writing to file: " + str(err))
        return self

    def write_to_h5ad(self, filepath: str | None=None) -> "SpatialHandler":
        """
        Writes the current AnnData object (`self.adata`) to an H5AD file at the specified file path.

        Parameters:
            filepath (str | None): The destination file path where the AnnData object should be written.
                If None, an exception is raised.

        Returns:
            SpatialHandler: Returns self after successfully writing the file.

        Raises:
            Exception: If no AnnData object is present, if no file path is provided, or if an error occurs during writing.
                In case of a write error, any partially written file at the destination path is removed.
        """
        if self.adata is None:
            raise Exception("No AnnData object present to write to file.")
        if filepath is None:
            raise Exception("No destination file path given. Provide one to write file.")
        try:
            self.adata.write(filename=Path(filepath))
        except Exception as err:
            # remove the file if it was created
            if os.path.exists(filepath):
                os.remove(filepath)
            raise Exception("Error occurred while writing to file: ", err)
        return self

class CoxMxHandler(SpatialHandler):
    """
    Factory class for CoxMx dataset uploads and conversions.

    Standardized names for different files:
    * <dataset_id>_`'anndata.h5ad'`: Counts and metadata file.
    * <dataset_id>_`'cluster_assignment.txt'`: Cluster assignment file.
    * <dataset_id>_`'Metrics.csv'`: Metrics file.
    * <dataset_id>_`'variable_features_clusters.txt'`: Variable features clusters file.
    * <dataset_id>_`'variable_features_spatial_moransi.txt'`: Variable features Moran’s I file.
    """

    @property
    def has_images(self) -> bool:
        """Whether this handler has associated images (always False for CoxMx)."""
        return False

    @property
    def coordinate_system(self) -> str:
        """Returns the coordinate system used by CoxMx datasets."""
        return "global"

    @property
    def region_id(self) -> str:
        """Returns the region ID used for spot data."""
        return "instance_id"

    @property
    def region_name(self) -> str:
        """Returns the name of the region used for spot data."""
        return "locations"

    @property
    def platform(self) -> str:
        """Returns the platform name for this handler."""
        return "coxmx"

    @property
    def img_name(self) -> str | None:
        """Returns the image name associated with this handler (always None for CoxMx)."""
        return None

    def process_file(self, filepath: str, **kwargs) -> "SpatialHandler":
        """
        Reads and processes a CoxMx spatial data file from the given filepath.
        For CoxMx, this is a stub and does not perform any operation.
        """
        return self


class CurioHandler(SpatialHandler):
    """
    Factory class for Curio Seeker dataset uploads and conversions.

    Standardized names for different files:
    * <dataset_id>_`'anndata.h5ad'`: Counts and metadata file.
    * <dataset_id>_`'cluster_assignment.txt'`: Cluster assignment file.
    * <dataset_id>_`'Metrics.csv'`: Metrics file.
    * <dataset_id>_`'variable_features_clusters.txt'`: Variable features clusters file.
    * <dataset_id>_`'variable_features_spatial_moransi.txt'`: Variable features Moran’s I file.

    It seems Curio Seeker only provides gene IDs (as indexes to an empty dataframe),
    so you need to modify the included h5ad file to include ensembl IDs in adata.var.
    You can use add_ensembl_id_to_h5ad_missing_release.py for that, and include the revisted h5ad file in the tarball.
    """

    @property
    def has_images(self) -> bool:
        """Whether this handler has associated images (always False for Curio)."""
        return False

    @property
    def coordinate_system(self) -> str:
        """Returns the coordinate system used by Curio datasets."""
        return "global"

    @property
    def region_id(self) -> str:
        """Returns the region ID used for spot data."""
        return "instance_id"

    @property
    def region_name(self) -> str:
        """Returns the name of the region used for spot data."""
        return "cells"

    @property
    def platform(self) -> str:
        """Returns the platform name for this handler."""
        return "curio"

    @property
    def img_name(self) -> str | None:
        """Returns the image name associated with this handler (always None for Curio)."""
        return None

    def process_file(self, filepath: str, **kwargs) -> "SpatialHandler":
        """
        Reads and processes a Curio Seeker spatial data tarball from the given filepath.
        Extracts and processes the h5ad and Moran's I-score files, updates gene IDs, and loads into a SpatialData object.
        """
        extract_dir = kwargs.get("extract_dir", '/tmp/')
        extract_dir = os.path.join(extract_dir, 'files')

        h5ad_file = None
        spatial_moransi_file = None

        if filepath.endswith(".tar.gz"):
            mode = "r:gz"  # Read as gzipped tar file
        elif filepath.endswith(".tar"):
            mode = "r"     # Read as plain tar file
        else:
            raise Exception("File must be a .tar or .tar.gz file.")

        if os.path.isdir(extract_dir):
            # Remove any existing directory
            os.system("rm -rf {}".format(extract_dir))

        with tarfile.open(filepath, mode) as tf:
            for entry in tf:
                # Skip any BSD tar artifacts, like files that start with ._ or .DS_Store
                if ".DS_Store" in entry.name or "._" in entry.name:
                    continue
                # Extract file into tmp dir
                filepath = "{0}/{1}".format(extract_dir, entry.name)
                tf.extract(entry, path=extract_dir)

                if entry.name.endswith("anndata.h5ad"):
                    h5ad_file = filepath
                elif entry.name.endswith("variable_features_spatial_moransi.txt"):
                    spatial_moransi_file = filepath

        """
        The sdio.curio function below uses the Moran's I-score file to overwrite the adata.var that was previously set
        with gene IDs. We need to pull out the original h5ad file and modify the Moran's I-score file to use the ensembl IDs.
        """

        if h5ad_file is None:
            raise Exception("h5ad file not found in tarball.")

        # Try to get organism id directly or through dataset metadata
        organism_id = kwargs.get("organism_id", None)
        if organism_id is None and "dataset_id" in kwargs:
            from geardb import get_dataset_by_id
            dataset = get_dataset_by_id(kwargs.get("dataset_id"))   # assumes the metadata is already present
            if dataset:
                organism_id = dataset.organism_id
        if organism_id is None:
            raise Exception("Organism ID not found in dataset metadata or provided as an argument.")

        # Read in the h5ad file
        adata = ad.read_h5ad(h5ad_file)

        # Add ensemble IDs to the adata.var
        adata = update_adata_with_ensembl_ids(adata, organism_id, "UNMAPPED_")

        # Create mapping dict.
        # Original index name (ensembl_id) was created in add_ensembl_id_to_h5ad_missing_release.py
        gene_symbol_to_ensembl_id = adata.var.reset_index().set_index("gene_symbol").to_dict()["ensembl_id"]
        if spatial_moransi_file is None:
            raise Exception("Moran's I score file not found in tarball.")

        var_features_moransi = pd.read_csv(spatial_moransi_file, sep="\t", header=0)

        # Replace the gene symbols with the ensembl ids
        var_features_moransi.index = var_features_moransi.index.map(gene_symbol_to_ensembl_id)
        var_features_moransi.to_csv(spatial_moransi_file, sep="\t", header=True, index=True, index_label=False)

        # Now are ready to read in to a SpatialData object
        sdata = sdio.curio(extract_dir)

        # To get the adata equivalent, look at sdata.tables["table"]

        # The Space Ranger h5 matrix has the gene names as the index, need to move them to a column and set the index to the ensembl id
        sdata.tables[self.NORMALIZED_TABLE_NAME].var_names_make_unique()

        # Change obs "cluster" to "clusters" to harmonize
        sdata.tables[self.NORMALIZED_TABLE_NAME].obs = sdata.tables[self.NORMALIZED_TABLE_NAME].obs.rename(columns={"cluster": "clusters"})

        self.sdata = sdata
        self.standardize_sdata()

        # table name should already be "table" for Visium

        self.originalFile = filepath
        return self

class GeoMxHandler(SpatialHandler):
    """
    Code is mostly inspired by https://github.com/LiHongCSBLab/SOAPy/blob/153095a44200a07a73a6a72c9978adfa1581c853/SOAPy_st/pp/all2adata.py#L229
    I wanted to install SOAPy but ran into pip requirement compatibility issues.  For example, we use a later version of AnnData in gEAR than SOAPy does.

    Factory class for GeoMx dataset uploads and conversions.

    Required files:
    * "xlsx" file with information.
      * This Excel file must contain a sheet named "SegmentProperties" with a column named "SegmentDisplayName" which will be used as the cell ID.
      * This Excel file must contain a sheet named "TargetCountMatrix" or "BioProbeCountMatrix" with the counts matrix.

    NOT IMPLEMENTED - Polygon data from XML files

    It seems GeoMx Excel data has gene IDs but can have multiple accessions (none are ensembl IDs),
    so you need to modify add the ensembl IDs to the adata.var.
    You can use add_ensembl_id_to_h5ad_missing_release.py for that, and include the revisted h5ad file in the tarball.

    """

    @property
    def has_images(self) -> bool:
        """Whether this handler has associated images (always False for GeoMx)."""
        return False

    @property
    def coordinate_system(self) -> str:
        """Returns the coordinate system used by GeoMx datasets."""
        return "global"

    @property
    def region_id(self) -> str:
        """Returns the region ID used for spot data."""
        return "instance_id"

    @property
    def region_name(self) -> str:
        """Returns the name of the region used for spot data."""
        return "locations"

    @property
    def platform(self) -> str:
        """Returns the platform name for this handler."""
        return "geomx"

    @property
    def img_name(self) -> str | None:
        """Returns the image name associated with this handler (always None for GeoMx)."""
        return None

    def process_file(self, filepath: str, **kwargs) -> "SpatialHandler":
        """
        Reads and processes a GeoMx spatial data tarball from the given filepath.
        Extracts the Excel file, validates required sheets, loads counts and metadata, updates gene IDs, and loads into a SpatialData object.
        """
        extract_dir = kwargs.get("extract_dir", '/tmp/')
        extract_dir = os.path.join(extract_dir, 'files')

        if filepath.endswith(".tar.gz"):
            mode = "r:gz"  # Read as gzipped tar file
        elif filepath.endswith(".tar"):
            mode = "r"     # Read as plain tar file
        else:
            raise Exception("File must be a .tar or .tar.gz file.")

        if os.path.isdir(extract_dir):
            # Remove any existing directory
            os.system("rm -rf {}".format(extract_dir))

        information_file = None

        with tarfile.open(filepath, mode) as tf:
            for entry in tf:
                # Skip any BSD tar artifacts, like files that start with ._ or .DS_Store
                if ".DS_Store" in entry.name or "._" in entry.name:
                    continue
                # Extract file into tmp dir
                filepath = "{0}/{1}".format(extract_dir, entry.name)
                tf.extract(entry, path=extract_dir)

                if entry.name.endswith(".xlsx"):
                    information_file = filepath

        if information_file is None:
            raise Exception("Excel file containing sample and count information not found in tarball.")

        # Validate the Excel file
        import openpyxl
        wb = openpyxl.load_workbook(information_file)
        if "SegmentProperties" not in wb.sheetnames:
            raise Exception("Excel file must contain a sheet named 'SegmentProperties' with a column named 'SegmentDisplayName'.")
        # Determine if the count matrix is in "TargetCountMatrix" or "BioProbeCountMatrix" and store sheet name as variable
        if "TargetCountMatrix" in wb.sheetnames:
            count_matrix_sheet = "TargetCountMatrix"
        elif "BioProbeCountMatrix" in wb.sheetnames:
            count_matrix_sheet = "BioProbeCountMatrix"
        else:
            raise Exception("Excel file must contain a sheet named 'TargetCountMatrix' or 'BioProbeCountMatrix' with the counts matrix.")

        # Try to get organism id directly or through dataset metadata
        organism_id = kwargs.get("organism_id", None)
        if organism_id is None and "dataset_id" in kwargs:
            from geardb import get_dataset_by_id
            dataset = get_dataset_by_id(kwargs.get("dataset_id"))   # assumes the metadata is already present
            if dataset:
                organism_id = dataset.organism_id
        if organism_id is None:
            raise Exception("Organism ID not found in dataset metadata, sample taxon id, or provided as an argument.")

        # Get observation table data
        roi_obs_df = pd.read_excel(
            information_file,
            sheet_name='SegmentProperties',
            index_col='SegmentDisplayName',
            header=0,
        )
        obs = roi_obs_df

        # Get count matrix data
        roi_counts_df = pd.read_excel(
            information_file,
            sheet_name=count_matrix_sheet,
            index_col=0,
            header=0,
        ).T

        if count_matrix_sheet == "BioProbeCountMatrix":
            # If the count matrix is from BioProbeCountMatrix, need to set TargetName values as the column names
            # The sheet has accession IDs but some entries have multiple IDs, so we will map Ensembl IDs later
            roi_counts_df.columns = roi_counts_df.loc["TargetName"].to_numpy()
            roi_counts_df = roi_counts_df.drop("TargetName")

        counts = roi_counts_df.loc[obs.index, :]

        var = pd.DataFrame(index=counts.columns)

        adata = ad.AnnData(counts.values, obs=obs, var=var, uns={}, obsm={})
        adata.obsm['spatial'] = obs.loc[:, ['ROICoordinateX', 'ROICoordinateY']].to_numpy()

        # Sanitize adata.obs so that column names only contain alphanumeric characters, underscores, dots and hyphens.
        adata.obs.columns = adata.obs.columns.str.replace(r'+', '_Plus')    # special case for '+' character
        adata.obs.columns = adata.obs.columns.str.replace(r'[^a-zA-Z0-9_.-]', '_')
        adata.obs.columns = adata.obs.columns.str.replace(r' ', '_')

        # Add ensemble IDs to the adata.var
        adata = update_adata_with_ensembl_ids(adata, organism_id, "UNMAPPED_")

        # Convert to SpatialData object
        sdata = from_legacy_anndata(adata)

        # GeoMx is focused more on spatial bulk rather than spatial single-cell, so we will set the clusters to the index
        sdata.tables[self.NORMALIZED_TABLE_NAME].obs["clusters"] = sdata.tables[self.NORMALIZED_TABLE_NAME].obs.index

        # Have the shapes "locations" df index match the obs index
        sdata.shapes["locations"].index = sdata.tables[self.NORMALIZED_TABLE_NAME].obs[self.region_id]

        self.sdata = sdata
        self.standardize_sdata()
        self.originalFile = filepath
        return self

class VisiumHandler(SpatialHandler):
    # NOTE: Uploads work but it cannot be used in a spatial panel yet because clusters have not been provided.
    """
    Factory class for Visium dataset uploads and conversions.

    Standardized names for different files:
    * (<dataset_id>_)`'filtered_feature_bc_matrix.h5'`: Counts and metadata file.
    * 'analysis/clustering/gene_expression_graphclust/clusters.csv': Clustering information. Preferable if "Cluster" column has actual labels instead of numbers.
    * 'spatial/tissue_hires_image.png': High resolution image.
    * 'spatial/tissue_lowres_image.png': Low resolution image.
    * 'spatial/scalefactors_json.json': Scalefactors file.
    * 'spatial/tissue_positions_list.csv' (SpaceRanger 1) or 'spatial/tissue_positions.csv' (SpaceRanger 2): Spots positions file.
    * fullres_image_file: (NOT USED) large microscopy image used as input for space ranger.

    Recommended tar command to create tarball:
    `tar cvf <dataset>.tar <dataset>_filtered_feature_bc_matrix.h5 spatial analysis`
    """

    @property
    def has_images(self) -> bool:
        """Whether this handler has associated images (always True for Visium)."""
        return True

    @property
    def coordinate_system(self) -> str:
        """Returns the coordinate system used by Visium datasets."""
        return "downscaled_hires"

    @property
    def region_id(self) -> str:
        """Returns the region ID used for spot data."""
        return "spot_id"

    @property
    def region_name(self) -> str:
        """Returns the name of the region used for spot data."""
        return "spatialdata"

    @property
    def platform(self) -> str:
        """Returns the platform name for this handler."""
        return "visium"

    @property
    def img_name(self) -> str | None:
        """Returns the image name associated with this handler."""
        return "spatialdata_hires_image"

    def process_file(self, filepath: str, **kwargs) -> "SpatialHandler":

        extract_dir = kwargs.get("extract_dir", '/tmp/')
        extract_dir = os.path.join(extract_dir, 'files')

        if filepath.endswith(".tar.gz"):
            mode = "r:gz"  # Read as gzipped tar file
        elif filepath.endswith(".tar"):
            mode = "r"     # Read as plain tar file
        else:
            raise Exception("File must be a .tar or .tar.gz file.")

        if os.path.isdir(extract_dir):
            # Remove any existing directory
            os.system("rm -rf {}".format(extract_dir))

        with tarfile.open(filepath, mode) as tf:
            for entry in tf:
                # Skip any BSD tar artifacts, like files that start with ._ or .DS_Store
                if ".DS_Store" in entry.name or "._" in entry.name:
                    continue
                # Extract file into tmp dir
                filepath = "{0}/{1}".format(extract_dir, entry.name)
                tf.extract(entry, path=extract_dir)


        clustering_csv_path = "{}/analysis/clustering/gene_expression_graphclust/clusters.csv".format(extract_dir)
        # If clustering file does not exist, raise an exception
        if not os.path.exists(clustering_csv_path):
            raise Exception("clusters.csv file not found in tarball.")

        # If clustering file does not have "Barcode" and "Cluster" columns, raise an exception
        with open(clustering_csv_path, 'r') as f:
            first_line = f.readline()
            if "Barcode" not in first_line or "Cluster" not in first_line:
                raise Exception("clusters.csv file does not have 'Barcode' and 'Cluster' columns in clusters.csv file in tarball.")

        sdata = sdio.visium(path=extract_dir, dataset_id="spatialdata")    # Provide a name to standarize downstream usage

        # add clustering information to the vis_sdata.table.obs dataframe
        clustering = pd.read_csv(clustering_csv_path)
        # make barcode as index
        clustering = clustering.set_index('Barcode')
        sdata.tables[self.NORMALIZED_TABLE_NAME].obs['clusters'] = clustering['Cluster'].astype('category')

        # The Space Ranger h5 matrix has the gene names as the index, need to move them to a column and set the index to the ensembl id
        sdata.tables[self.NORMALIZED_TABLE_NAME].var_names_make_unique()

        # currently gene symbols are the index, need to move them to a column
        sdata.tables[self.NORMALIZED_TABLE_NAME].var["gene_symbol"] = sdata.tables[self.NORMALIZED_TABLE_NAME].var.index

        # set the index to the ensembl id (gene_ids)
        sdata.tables[self.NORMALIZED_TABLE_NAME].var = sdata.tables[self.NORMALIZED_TABLE_NAME].var.set_index("gene_ids")

        self.sdata = sdata
        self.standardize_sdata()
        self.originalFile = filepath
        return self

class VisiumHDHandler(SpatialHandler):
    """
    Factory class for Visium HD dataset uploads and conversions.

    Explanation of Space Ranger v3 output here -> https://www.10xgenomics.com/support/software/space-ranger/latest/analysis/outputs/output-overview#hd-outputs

    Required files that will break the upload if not present:
    /binned_outputs/square_008um/filtered_feature_bc_matrix.h5
    /binned_outputs/square_008um/analysis/clustering/gene_expression_graphclust/clusters.csv
    /binned_outputs/feature_slice.h5
    /binned_outputs/square_008um/spatial/scalefactors_json.json
    /binned_outputs/square_008um/spatial/tissue_hires_image.png
    /binned_outputs/square_008um/spatial/tissue_lowres_image.png

    Recommended tar command to create tarball:
    `tar cvf <dataset>.tar binned_outputs/feature_slice.h5 binned_outputs/square_008um`

    Special note: We have observed that bin sizes finer than 8 microns per pixel will generally have more cells, which can lead to longer analysis times.
    For now, we will attempt to use the "square_008um" binned output.

    """

    table_name = "square_008um"

    @property
    def has_images(self) -> bool:
        """Whether this handler has associated images (always True for Visium HD)."""
        return True

    @property
    def coordinate_system(self) -> str:
        """Returns the coordinate system used by Visium HD datasets."""
        return "downscaled_hires"

    @property
    def region_id(self) -> str:
        """Returns the region ID used for spot data."""
        return "location_id"

    @property
    def region_name(self) -> str:
        """Returns the name of the region used for spot data."""
        return "spatialdata_square_008um"

    @property
    def platform(self) -> str:
        """Returns the platform name for this handler."""
        return "visium_hd"

    @property
    def img_name(self) -> str | None:
        """Returns the image name associated with this handler."""
        return "spatialdata_hires_image"

    def process_file(self, filepath: str, **kwargs) -> "SpatialHandler":
        extract_dir = kwargs.get("extract_dir", '/tmp/')
        extract_dir = os.path.join(extract_dir, 'files')

        binned_outputs_dir = "{}/binned_outputs".format(extract_dir)
        bin_008_dataset_path = "{}/{}/".format(binned_outputs_dir, self.table_name)
        clustering_csv_path = "{}/analysis/clustering/gene_expression_graphclust/clusters.csv".format(bin_008_dataset_path)

        absolute_path = os.path.abspath(binned_outputs_dir)

        if filepath.endswith(".tar.gz"):
            mode = "r:gz"  # Read as gzipped tar file
        elif filepath.endswith(".tar"):
            mode = "r"     # Read as plain tar file
        else:
            raise Exception("File must be a .tar or .tar.gz file.")

        if os.path.isdir(extract_dir):
            # Remove any existing directory
            os.system("rm -rf {}".format(extract_dir))

        with tarfile.open(filepath, mode) as tf:
            for entry in tf:
                # Skip any BSD tar artifacts, like files that start with ._ or .DS_Store
                if ".DS_Store" in entry.name or "._" in entry.name:
                    continue

                # IF directory has "square_" but not "square_008um", skip
                if "square_" in entry.name and "square_008um" not in entry.name:
                    continue

                # Extract file into tmp dir
                filepath = "{0}/{1}".format(extract_dir, entry.name)
                tf.extract(entry, path=extract_dir)

        if not os.path.exists("{}/feature_slice.h5".format(binned_outputs_dir)):
            raise Exception("feature_slice.h5 file not found in /binned_outputs directory in tarball.")

        # If clustering file does not exist, raise an exception
        if not os.path.exists(clustering_csv_path):
            raise Exception("clusters.csv file not found in tarball.")

        # If clustering file does not have "Barcode" and "Cluster" columns, raise an exception
        with open(clustering_csv_path, 'r') as f:
            first_line = f.readline()
            if "Barcode" not in first_line or "Cluster" not in first_line:
                raise Exception("clusters.csv file does not have 'Barcode' and 'Cluster' columns in clusters.csv file in tarball.")

        # https://github.com/scverse/spatialdata-io/issues/212

        # Create a symlink for feature_slice.h5 within "visium_dataset_path" to include a "spatialdata_" prefix
        # This is a workaround for the current implementation of the visium_hd function

        if not os.path.exists("{}/spatialdata_feature_slice.h5".format(binned_outputs_dir)):
            os.symlink("{}/feature_slice.h5".format(absolute_path), "{}/spatialdata_feature_slice.h5".format(binned_outputs_dir))

        sdata = sdio.visium_hd(binned_outputs_dir
                                , dataset_id="spatialdata"   # Provide a name to standarize downstream usage
                                , bin_size=8
                                , filtered_counts_file=True
                                , load_all_images=False  # CytAssist image is not helpful for us.
                                , fullres_image_file=None
                                , bins_as_squares=True
                                )

        # add clustering information to the vis_sdata.table.obs dataframe
        clustering = pd.read_csv(clustering_csv_path)
        # make barcode as index
        clustering = clustering.set_index('Barcode')
        sdata.tables[self.table_name].obs['clusters'] = clustering['Cluster'].astype('category')

        # To get the adata equivalent, look at sdata.tables["table"]
        # The Space Ranger h5 matrix has the gene names as the index, need to move them to a column and set the index to the ensembl id
        sdata.tables[self.table_name].var_names_make_unique()

        # currently gene symbols are the index, need to move them to a column
        sdata.tables[self.table_name].var["gene_symbol"] = sdata.tables[self.table_name].var.index

        # set the index to the ensembl id (gene_ids)
        sdata.tables[self.table_name].var = sdata.tables[self.table_name].var.set_index("gene_ids")

        # Set the table name to the normalized table name
        sdata.tables[self.NORMALIZED_TABLE_NAME] = sdata.tables[self.table_name]

        self.sdata = sdata
        self.standardize_sdata()
        self.originalFile = filepath
        return self

class XeniumHandler(SpatialHandler):
    """
    Factory class for Xenium dataset uploads and conversions.

    Standardized names for different files:
    * (REQ) 'experiment.xenium': File containing specifications.
    * (REQ) 'cell_feature_matrix.h5': File containing cell feature matrix.
    * (REQ) 'cells.parquet': File containing cell metadata.
    * (REQ) 'morphology_focus.ome.tif': File containing morphology focus or a "morphology_focus" directory containing multiple images.
    * 'nucleus_boundaries.parquet': Polygons of nucleus boundaries.
    * 'cell_boundaries.parquet': Polygons of cell boundaries.
    * 'transcripts.parquet': File containing transcripts.
    * 'cells.zarr.zip': Zarr file containing cell and nucleus label data (NOT USING FOR NOW)
    * 'analysis/clustering/gene_expression_graphclust/clusters.csv': Clustering information. Preferable if "Cluster" column has actual labels instead of numbers.

    Currently we are not using the cells.zarr.zip file because of memory issues when converting the resulting cell labels as part of the AnnData conversion process.

    More information on Xenium Ranger outputs can be found here: https://www.10xgenomics.com/support/software/xenium-ranger/latest/analysis/outputs/XR-output-overview
    """

    @property
    def has_images(self) -> bool:
        """Whether this handler has associated images (always True for Xenium)."""
        return True

    @property
    def coordinate_system(self) -> str:
        """Returns the coordinate system used by Xenium datasets."""
        return "global"

    @property
    def region_id(self) -> str:
        """Returns the region ID used for spot data."""
        return "cell_id"

    @property
    def region_name(self) -> str:
        """Returns the name of the region used for spot data."""
        return "cell_circles"

    @property
    def platform(self) -> str:
        """Returns the platform name for this handler."""
        return "xenium"

    @property
    def img_name(self) -> str | None:
        """Returns the image name associated with this handler."""
        return "morphology_focus"

    def process_file(self, filepath: str, **kwargs) -> "SpatialHandler":
        """
        Reads and processes a Xenium spatial data tarball from the given filepath.
        Extracts required files, loads clustering and spatial data, updates gene IDs, and loads into a SpatialData object.
        """
        extract_dir = kwargs.get("extract_dir", '/tmp/')
        extract_dir = os.path.join(extract_dir, 'files')

        if filepath.endswith(".tar.gz"):
            mode = "r:gz"  # Read as gzipped tar file
        elif filepath.endswith(".tar"):
            mode = "r"     # Read as plain tar file
        else:
            raise Exception("File must be a .tar or .tar.gz file.")

        if os.path.isdir(extract_dir):
            # Remove any existing directory
            os.system("rm -rf {}".format(extract_dir))

        # settings to enable or disable based on if a file is present in the uploaded tarball
        include_raster_labels = False
        cell_boundaries_present = False
        nucleus_boundaries_present = False
        transcripts_present = False

        with tarfile.open(filepath, mode) as tf:
            for entry in tf:
                # Skip any BSD tar artifacts, like files that start with ._ or .DS_Store
                if ".DS_Store" in entry.name or "._" in entry.name:
                    continue
                # Extract file into tmp dir
                filepath = "{0}/{1}".format(extract_dir, entry.name)
                tf.extract(entry, path=extract_dir)

                if entry.name == "cells.zarr.zip":
                    include_raster_labels = True
                if entry.name == "cell_boundaries.parquet":
                    cell_boundaries_present = True
                if entry.name == "nucleus_boundaries.parquet":
                    nucleus_boundaries_present = True
                if entry.name == "transcripts.parquet":
                    transcripts_present = True

        # If clustering file does not exist, raise an exception
        clustering_csv_path = "{}/analysis/clustering/gene_expression_graphclust/clusters.csv".format(extract_dir)
        if not os.path.exists(clustering_csv_path):
            raise Exception("clusters.csv file not found in tarball.")

        # If clustering file does not have "Barcode" and "Cluster" columns, raise an exception
        with open(clustering_csv_path, 'r') as f:
            first_line = f.readline()
            if "Barcode" not in first_line or "Cluster" not in first_line:
                raise Exception("clusters.csv file does not have 'Barcode' and 'Cluster' columns in clusters.csv file in tarball.")

        sdata = sdio.xenium(extract_dir
                            , cells_labels=False # Avoid adding polygons to SpatialData object (for now due to out-of-memory issues)
                            , nucleus_labels=False
                            , cell_boundaries=cell_boundaries_present
                            , nucleus_boundaries=nucleus_boundaries_present
                            , transcripts=transcripts_present
                            , cells_as_circles=True  # Table is associated with the cells instead of the nuclei (faster performance)
                            , morphology_mip=False   # Using the morphology_focus image instead
                            )

        # In code, it seems that the Xenium reader is supposed to set the index to the "barcodes" column
        # But this column is not found, so we need to manually replace with "cell_id"
        sdata.tables[self.NORMALIZED_TABLE_NAME].obs["Barcode"] = sdata.tables[self.NORMALIZED_TABLE_NAME].obs["cell_id"]
        sdata.tables[self.NORMALIZED_TABLE_NAME].obs = sdata.tables[self.NORMALIZED_TABLE_NAME].obs.set_index("Barcode")

        # Change annotation target from "cell_circles" to "cell_labels"
        #sdata["table"].obs["region"] = "cell_labels"
        #sdata.set_table_annotates_spatialelement(
        #    table_name="table", region="cell_labels", region_key="region", instance_key="cell_labels"
        #)

        # add clustering information to the vis_sdata.table.obs dataframe
        clustering = pd.read_csv(clustering_csv_path)
        # make barcode as index
        clustering = clustering.set_index('Barcode')
        sdata.tables[self.NORMALIZED_TABLE_NAME].obs['clusters'] = clustering['Cluster'].astype('category')

        # The Space Ranger h5 matrix has the gene names as the index, need to move them to a column and set the index to the ensembl id
        sdata.tables[self.NORMALIZED_TABLE_NAME].var_names_make_unique()

        # currently gene symbols are the index, need to move them to a column
        sdata.tables[self.NORMALIZED_TABLE_NAME].var["gene_symbol"] = sdata.tables[self.NORMALIZED_TABLE_NAME].var.index

        # set the index to the ensembl id (gene_ids)
        sdata.tables[self.NORMALIZED_TABLE_NAME].var = sdata.tables[self.NORMALIZED_TABLE_NAME].var.set_index("gene_ids")

        self.sdata = sdata
        self.standardize_sdata()
        self.originalFile = filepath
        return self


### Helper constants

SPATIALTYPE2CLASS = {
    #"cosmx": CoxMxHandler,
    "curio": CurioHandler,
    "geomx": GeoMxHandler,
    "visium": VisiumHandler,
    "visium_hd": VisiumHDHandler,
    "visiumhd": VisiumHDHandler,    # allow both with and without underscore
    "xenium": XeniumHandler
}

ORG_ID_REQ_TYPES = ["curio", "geomx"]